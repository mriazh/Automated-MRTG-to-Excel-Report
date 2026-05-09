import os
import sys

# ==============================================================================
# 🔇 THE ULTIMATE "KIPAS ANGIN" SILENCE (LOW-LEVEL DUPLICATION)
# ==============================================================================
_stdout_fd = sys.stdout.fileno()
_stderr_fd = sys.stderr.fileno()
_save_stdout = os.dup(_stdout_fd)
_save_stderr = os.dup(_stderr_fd)

with open(os.devnull, 'w') as fnull:
    # Bajak jalur pipa di level OS
    os.dup2(fnull.fileno(), _stdout_fd)
    os.dup2(fnull.fileno(), _stderr_fd)
    
    try:
        import warnings
        import re
        import io
        import logging
        import traceback
        import contextlib
        
        # Environment Flags
        os.environ['GLOG_minloglevel'] = '3'
        os.environ['FLAGS_minloglevel'] = '3'
        os.environ['PADDLE_LOG_LEVEL'] = 'ERROR'
        os.environ['PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK'] = 'True'
        os.environ['PADDLEX_DISABLE_PRINT'] = '1'
        os.environ['FLAGS_enable_pir_api'] = '0'
        os.environ['FLAGS_enable_new_executor'] = '0' 
        os.environ['FLAGS_use_onednn'] = '0' 
        os.environ['FLAGS_use_mkldnn'] = '0'
        os.environ['FLAGS_use_gpu'] = '0'
        os.environ['KMP_WARNINGS'] = '0'

        # Mute warnings
        warnings.filterwarnings("ignore")
        
        # Silent Engine Import
        import paddle
        paddle.enable_static()
        from paddleocr import PaddleOCR
    except:
        pass
    finally:
        # Kembalikan jalur pipa asli
        os.dup2(_save_stdout, _stdout_fd)
        os.dup2(_save_stderr, _stderr_fd)
        os.close(_save_stdout)
        os.close(_save_stderr)
# ==============================================================================

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import column_index_from_string, get_column_letter
from PIL import Image as PILImage
from tqdm import tqdm
from rich.console import Console
from rich.style import Style

# ==============================================================================
# ⚙️ KONFIGURASI UTAMA (MAGIC NUMBERS)
# ==============================================================================
# --- Folder & File ---
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
FOLDER_DATA = os.path.join(BASE_DIR, "MRTG-Data")
LOG_FILE    = 'ocr_report.log'

# --- Tampilan Excel ---
IMAGE_SCALE         = 0.98    # Skala gambar MRTG di dalam sel (0.1 - 1.0)
EXCEL_DEFAULT_WIDTH = 8.43    # Lebar kolom standar Excel
EXCEL_DEFAULT_HEIGHT = 15.0    # Tinggi baris standar Excel (points)
PX_CONV_WIDTH       = 7.4     # Faktor konversi Lebar Kolom -> Pixels
PX_CONV_HEIGHT      = 1.333   # Faktor konversi Tinggi Baris -> Pixels

# --- Tampilan Terminal (UI) ---
TERMINAL_WIDTH      = 60      # Lebar garis pemisah (====)
PBAR_WIDTH_OCR      = 115     # Lebar progress bar mode OCR
PBAR_WIDTH_IMG      = 100     # Lebar progress bar mode Image Only

# --- Path Template & Output (OCR) ---
OCR_TEMPLATE = os.path.join(BASE_DIR, "MRTG-Monthly-Report-on-Internet-Bandwidth-Utilization-by-Telkom.xlsx")
OCR_OUTPUT   = os.path.join(BASE_DIR, "MRTG-Monthly-Report.xlsx")
OCR_MAPPING  = os.path.join(BASE_DIR, "list_mrtg_data_position.txt")
OCR_DAFTAR   = os.path.join(BASE_DIR, "list_mrtg_data.txt")

# --- Path Template & Output (Image Only) ---
IMG_TEMPLATE = os.path.join(BASE_DIR, "MRTG-Monthly-Report-on-Internet-Bandwidth-Utilization-by-Telkom (Img only).xlsx")
IMG_OUTPUT   = os.path.join(BASE_DIR, "MRTG-Monthly-Report-image-only.xlsx")
IMG_MAPPING  = os.path.join(BASE_DIR, "list_mrtg_data_position_img_only.txt")
IMG_DAFTAR   = os.path.join(BASE_DIR, "list_mrtg_data_img_only.txt")
# ==============================================================================

# Rich Console untuk output yang cantik
from rich.console import Console
console = Console()


# ==============================================================================
# 🔇 LOGGING SETUP (DEBUG to File, INFO to Console)
# ==============================================================================
# Setup Root Logger
root_logger = logging.getLogger()
root_logger.setLevel(logging.DEBUG)

# Clear existing handlers
for handler in root_logger.handlers[:]:
    root_logger.removeHandler(handler)

# 1. File Handler (Detail: DEBUG)
file_handler = logging.FileHandler(LOG_FILE, encoding='utf-8')
file_handler.setLevel(logging.DEBUG)
file_formatter = logging.Formatter('[%(asctime)s] [%(levelname)8s] %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
file_handler.setFormatter(file_formatter)
root_logger.addHandler(file_handler)

# 2. Console Handler (Clean: INFO)
console_handler = logging.StreamHandler(sys.stdout)
console_handler.setLevel(logging.INFO)
console_formatter = logging.Formatter('%(message)s')
console_handler.setFormatter(console_formatter)
root_logger.addHandler(console_handler)

# Final logger for this script
logger = logging.getLogger('mrtg_report')

# Matikan log internal library sebisanya
for name in ["ppocr", "paddlex", "ppstructure", "paddle", "PIL", "urllib3"]:
    l = logging.getLogger(name)
    l.setLevel(logging.ERROR)
    l.propagate = False
# ==============================================================================

def baca_daftar(filepath):
    """Baca daftar SID / Graph-title dari file teks."""
    items = []
    with open(filepath, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = line.split('.', 1)
            if len(parts) != 2:
                continue
            nomor = parts[0].strip()
            rest = parts[1].strip()
            if rest.startswith('SID : '):
                tipe = 'SID'
                id_val = rest.replace('SID : ', '').strip()
            elif rest.startswith('Graph-title : '):
                tipe = 'Graph-title'
                id_val = rest.replace('Graph-title : ', '').strip()
            else:
                continue
            items.append((nomor, tipe, id_val))
    return items


def get_area_size_pixels(sheet, start_row, start_col, end_row, end_col):
    """Hitung ukuran area Excel dalam pixel berdasarkan lebar kolom dan tinggi baris."""
    total_width = 0
    for col in range(start_col, end_col + 1):
        col_letter = get_column_letter(col)
        col_width = sheet.column_dimensions[col_letter].width
        if col_width is None:
            col_width = EXCEL_DEFAULT_WIDTH
        total_width += col_width * PX_CONV_WIDTH

    total_height = 0
    for row in range(start_row, end_row + 1):
        row_height = sheet.row_dimensions[row].height
        if row_height is None:
            row_height = EXCEL_DEFAULT_HEIGHT
        total_height += row_height * PX_CONV_HEIGHT

    return total_width, total_height


def resize_image_stretch(image_path, target_width, target_height):
    """Resize gambar ke ukuran target (stretch)."""
    with PILImage.open(image_path) as img:
        if img.mode in ('RGBA', 'LA', 'P'):
            img = img.convert('RGB')
        img_resized = img.resize((int(target_width), int(target_height)), PILImage.Resampling.LANCZOS)
        output = io.BytesIO()
        img_resized.save(output, format='PNG')
        output.seek(0)
        return output


def cari_path_gambar(folder_data, tanggal_str, tipe, id_val):
    """Cari path gambar MRTG berdasarkan tipe (SID/Graph-title) dan ID."""
    if tipe == 'SID':
        nama_file = f"MRTG_{id_val}.png"
    else:
        nama_file = f"MRTG_{id_val}_{tanggal_str}.png"

    path_gambar = os.path.join(folder_data, tanggal_str, nama_file)
    if not os.path.exists(path_gambar):
        # Fallback: cari file dengan awalan MRTG_{id_val}
        folder_tgl = os.path.join(folder_data, tanggal_str)
        if os.path.exists(folder_tgl):
            for f in os.listdir(folder_tgl):
                if f.startswith(f"MRTG_{id_val}") and f.endswith(".png"):
                    path_gambar = os.path.join(folder_tgl, f)
                    break
    return path_gambar


def tambah_gambar_di_area(sheet, image_path, start_row, start_col, end_row, end_col, scale=IMAGE_SCALE):
    """Resize gambar sesuai area dan letakkan di sheet Excel."""
    try:
        width_px, height_px = get_area_size_pixels(sheet, start_row, start_col, end_row, end_col)
        width_px = width_px * scale
        height_px = height_px * scale
        img_bytes = resize_image_stretch(image_path, width_px, height_px)
        img = XLImage(img_bytes)
        img.anchor = f"{get_column_letter(start_col)}{start_row}"
        sheet.add_image(img)
        return True
    except Exception as e:
        logger.debug(f"Gagal tambah gambar: {e}")
        return False


def get_tanggal_list(folder_data):
    """Ambil daftar folder tanggal (format YYYYMMDD) dari folder data."""
    if not os.path.exists(folder_data):
        return []
    tanggal_list = [
        d for d in os.listdir(folder_data)
        if os.path.isdir(os.path.join(folder_data, d)) and d.isdigit() and len(d) == 8
    ]
    tanggal_list.sort()
    return tanggal_list


# ========================================================
#  MODE 1: OCR (Ekstrak data + insert gambar)
# ========================================================

def baca_mapping_ocr(filepath):
    """Baca file mapping OCR (format Service Id dengan 6 field Inbound/Outbound + Image)."""
    mapping = {}
    with open(filepath, 'r', encoding='utf-8') as f:
        lines = [line.strip() for line in f if line.strip()]
    i = 0
    while i < len(lines):
        line = lines[i]
        if line.startswith('Service Id :') or line.startswith('Service Id :'):
            id_raw = line.replace('Service Id :', '').replace('Service Id :', '').strip()
            id_clean = re.sub(r'^\(\d+\)\s*', '', id_raw)
            i += 1
            entry = {}
            for _ in range(6):
                if i >= len(lines):
                    break
                key_val = lines[i].split(':', 1)
                if len(key_val) == 2:
                    key = key_val[0].strip()
                    val = key_val[1].strip()
                    match = re.match(r'([A-Z]+)(\d+)', val)
                    if match:
                        col_letter = match.group(1)
                        row = int(match.group(2))
                        col = column_index_from_string(col_letter)
                        entry[key] = (row, col)
                i += 1
            if i < len(lines) and lines[i].startswith('Image :'):
                range_str = lines[i].replace('Image :', '').strip()
                start, end = range_str.split('-')
                start_col = column_index_from_string(re.match(r'[A-Z]+', start).group())
                start_row = int(re.search(r'\d+', start).group())
                end_col = column_index_from_string(re.match(r'[A-Z]+', end).group())
                end_row = int(re.search(r'\d+', end).group())
                entry['Image'] = ((start_row, start_col), (end_row, end_col))
                i += 1
            mapping[id_clean] = entry
        else:
            i += 1
    return mapping


def _get_ocr_engine():
    """Singleton: inisialisasi PaddleOCR sekali saja."""
    if not hasattr(_get_ocr_engine, '_engine'):
        import warnings
        import contextlib
        import logging as py_logging
        
        # Mute internal Paddle loggers
        py_logging.getLogger("ppocr").setLevel(py_logging.ERROR)
        
        try:
            # Double check configuration
            paddle.set_flags({
                'FLAGS_enable_pir_api': 0,
                'FLAGS_enable_new_executor': 0,
                'FLAGS_use_onednn': 0
            })
        except:
            pass
            
        # Inisialisasi model super simpel biar nggak bentrok antar versi
        _get_ocr_engine._engine = PaddleOCR(lang='en')
            
        logger.debug("PaddleOCR engine siap.")
    return _get_ocr_engine._engine


def extract_mrtg_values(image_path):
    """Ekstrak nilai bandwidth dengan strategi Sequential Context & Fuzzy Keywords."""
    try:
        ocr = _get_ocr_engine()
        # Log ke file saja (DEBUG)
        logger.debug(f"OCR memproses: {os.path.basename(image_path)}")

        # Strategi panggil: Utamakan predict()
        if hasattr(ocr, 'predict'):
            result_iter = ocr.predict(image_path)
        elif hasattr(ocr, 'ocr'):
            result_iter = ocr.ocr(image_path)
        else:
            return None

        all_texts = []
        for res in result_iter:
            # Handle format output PaddleOCR vs PaddleX
            data = res.get('res', res) if isinstance(res, dict) else res
            if isinstance(data, list):
                # Format [[box, (text, conf)], ...]
                for line in data:
                    if isinstance(line, list) and len(line) > 1:
                        text = line[1][0] if isinstance(line[1], tuple) else str(line[1])
                        all_texts.append(str(text).strip())
            elif isinstance(data, dict):
                texts = data.get('rec_texts', [])
                all_texts.extend([str(t).strip() for t in texts])

        # --- DEBUG LOG: START ---
        sid_log = os.path.basename(image_path).replace('MRTG_', '').replace('.png', '')
        logger.debug(f"--- START OCR [{sid_log}] ---")
        logger.debug(f"Raw Text Detected: {all_texts}")

        # --- Helper: Cari nilai setelah keyword ---
        def find_value_after(keyword_list, texts, start_search_idx):
            """Mencari angka (+ unit) setelah salah satu keyword ditemukan."""
            for i in range(start_search_idx, len(texts)):
                t = texts[i].lower()
                # Cocokkan keyword dengan toleransi typo
                if any(kw.lower() in t for kw in keyword_list):
                    # Cari angka di i atau 3 box setelahnya
                    for j in range(i, min(i + 6, len(texts))):
                        # Pattern baru: Case-insensitive unit, dukung b/s, bps, dsb.
                        match = re.search(r'(\d+(?:[\.,]\d+)?)\s*([MkGTmkgt]?[Bb]?p?s?)', texts[j])
                        if match:
                            val = match.group(1).replace(',', '.')
                            unit_raw = match.group(2).strip()
                            
                            # Normalisasi Unit
                            unit = ""
                            u_low = unit_raw.lower()
                            if 't' in u_low: unit = "T"
                            elif 'g' in u_low: unit = "G"
                            elif 'm' in u_low: unit = "M"
                            elif 'k' in u_low: unit = "k"
                            
                            # Jika unit tidak nempel, cari di 2 box setelahnya
                            # Ini handle kasus PaddleOCR split nilai: '14.91' + 'M'
                            if not unit:
                                for k in range(j + 1, min(j + 3, len(texts))):
                                    next_t = texts[k].strip().lower()
                                    # Pastikan box berikutnya HANYA unit (pendek, tidak ada angka)
                                    if re.match(r'^[tmgkTMGK]$', next_t.strip()):
                                        if 't' in next_t: unit = "T"
                                        elif 'g' in next_t: unit = "G"
                                        elif 'm' in next_t: unit = "M"
                                        elif 'k' in next_t: unit = "k"
                                        break
                                    # Atau unit yang nempel tapi di box sendiri: 'M', 'k', dll
                                    elif next_t in ['m', 'k', 'g', 't']:
                                        if next_t == 't': unit = "T"
                                        elif next_t == 'g': unit = "G"
                                        elif next_t == 'm': unit = "M"
                                        elif next_t == 'k': unit = "k"
                                        break
                                    # Jika box berikutnya sudah berisi keyword lain, stop
                                    elif any(kw in next_t for kw in ['current', 'average', 'maximum', 'inbound', 'outbound', 'cur rent']):
                                        break
                            
                            # Default ke M jika masih kosong
                            if not unit:
                                unit = "M"
                                
                            return f"{val} {unit}".strip()
                        
                        if 'n/a' in texts[j].lower():
                            continue
            return "N/A"

        # --- Cari Section Inbound & Outbound dengan Fuzzy Matching ---
        inbound_idx = -1
        outbound_idx = -1
        
        # Keyword list yang lebih luas buat handle typo OCR
        in_kws = ['inbound', 'in-bound', 'in bound', 'inhound', '1nbound', 'nbound', 'inb']
        out_kws = ['outbound', 'out-bound', 'out bound', 'oulbound', '0utbound', 'outb']

        for i, t in enumerate(all_texts):
            t_low = t.lower()
            if any(kw in t_low for kw in in_kws):
                inbound_idx = i
            if any(kw in t_low for kw in out_kws):
                outbound_idx = i

        # Proteksi: Jika Outbound tidak ketemu di legend, cari penanda Current kedua
        if outbound_idx == -1 and inbound_idx != -1:
            for i in range(inbound_idx + 1, len(all_texts)):
                if any(kw in all_texts[i].lower() for kw in ['current', 'cur rent', 'cur ren', 'curren']):
                    # Pastikan ini Current untuk Outbound (setelah ada angka Inbound)
                    has_numbers_before = False
                    for j in range(inbound_idx + 1, i):
                        if re.search(r'\d', all_texts[j]):
                            has_numbers_before = True
                            break
                    if has_numbers_before:
                        outbound_idx = i - 1
                        break

        # Parse Inbound Area
        if inbound_idx != -1:
            search_limit = outbound_idx if outbound_idx > inbound_idx else len(all_texts)
            in_area = all_texts[inbound_idx:search_limit]
            in_vals = {
                'Current': find_value_after(['Current', 'Curren', 'Cur rent', 'Cur ren', 'Cur'], in_area, 0),
                'Average': find_value_after(['Average', 'Averaqe', 'Avera9e', 'Avera', 'Ave'], in_area, 0),
                'Maximum': find_value_after(['Maximum', 'Maximu', 'Maxlmu', 'Max'], in_area, 0)
            }
        else:
            in_vals = {'Current': 'N/A', 'Average': 'N/A', 'Maximum': 'N/A'}

        # Parse Outbound Area
        if outbound_idx != -1:
            out_area = all_texts[outbound_idx:]
            out_vals = {
                'Current': find_value_after(['Current', 'Curren', 'Cur rent', 'Cur ren', 'Cur'], out_area, 0),
                'Average': find_value_after(['Average', 'Averaqe', 'Avera9e', 'Avera', 'Ave'], out_area, 0),
                'Maximum': find_value_after(['Maximum', 'Maximu', 'Maxlmu', 'Max'], out_area, 0)
            }
        else:
            out_vals = {'Current': 'N/A', 'Average': 'N/A', 'Maximum': 'N/A'}

        result = {
            'Inbound_Current': in_vals['Current'],
            'Inbound_Average': in_vals['Average'],
            'Inbound_Maximum': in_vals['Maximum'],
            'Outbound_Current': out_vals['Current'],
            'Outbound_Average': out_vals['Average'],
            'Outbound_Maximum': out_vals['Maximum']
        }

        # --- DEBUG LOG: END ---
        logger.debug(f"Extracted Values: {result}")
        logger.debug(f"--- END OCR [{sid_log}] ---\n")

        return result
    except Exception as e:
        logger.error(f"OCR Critical Error [{os.path.basename(image_path)}]: {e}")
        return None


def tulis_nilai(sheet, entry, values):
    """Tulis nilai OCR ke sel Excel yang ditentukan."""
    for key, (row, col) in entry.items():
        if key in values and key != 'Image':
            sheet.cell(row=row, column=col, value=values[key])


def proses_tanggal_ocr(wb, tanggal_str, items, mapping, global_stats, review_list, tanggal_idx, total_tanggal, global_pbar):
    """Proses satu folder tanggal untuk mode OCR (ekstrak data + insert gambar)."""
    hari = int(tanggal_str[6:8])
    sheet_name_candidates = [str(hari), f"{hari:02d}"]
    sheet = None
    for name in sheet_name_candidates:
        if name in wb.sheetnames:
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb.create_sheet(title=str(hari))

    total_in_day = len(items)
    ok_count = 0
    na_count = 0
    fail_count = 0

    # Header pake ANSI color biar beneran berwarna di tqdm
    cyan = "\033[1;36m"
    dim = "\033[2m"
    reset = "\033[0m"
    global_pbar.write(f"\n{cyan}📅 TANGGAL: {tanggal_str} ({tanggal_idx}/{total_tanggal}){reset}")
    global_pbar.write(f"{dim}──────────────────────────────────────────────────{reset}")
    
    for idx, (nomor, tipe, id_val) in enumerate(items, 1):
        id_clean = re.sub(r'^\(\d+\)\s*', '', id_val)
        label = f"{id_val}"

        if id_clean not in mapping:
            global_pbar.write(f"  ⏭️  {label} (Skip: No Mapping)")
            global_pbar.update(1)
            continue
        entry = mapping[id_clean]

        path_gambar = cari_path_gambar(FOLDER_DATA, tanggal_str, tipe, id_val)
        if not os.path.exists(path_gambar):
            global_pbar.write(f"  ❌ {label} (Missing Image)")
            fail_count += 1
            global_stats['fail'] += 1
            review_list.append({'sid': id_clean, 'date': tanggal_str, 'sheet': sheet.title, 'status': 'Fail', 'na': 6})
            global_pbar.update(1)
            continue

        values = extract_mrtg_values(path_gambar)
        
        if not values:
            global_pbar.write(f"  ❌ {label} (OCR Fail)")
            fail_count += 1
            global_stats['fail'] += 1
            review_list.append({'sid': id_clean, 'date': tanggal_str, 'sheet': sheet.title, 'status': 'Fail', 'na': 6})
        else:
            # Hitung N/A
            val_na = sum(1 for v in values.values() if v == 'N/A')
            if val_na == 0:
                global_pbar.write(f"  ✅ {label} (Success)")
                ok_count += 1
                global_stats['ok'] += 1
            else:
                global_pbar.write(f"  ⚠️  {label} ({6-val_na}/6 OK)")
                na_count += 1
                global_stats['partial'] += 1
                review_list.append({'sid': id_clean, 'date': tanggal_str, 'sheet': sheet.title, 'status': 'Partial', 'na': val_na})

            tulis_nilai(sheet, entry, values)
            if 'Image' in entry:
                (start_row, start_col), (end_row, end_col) = entry['Image']
                tambah_gambar_di_area(sheet, path_gambar, start_row, start_col, end_row, end_col)

        # Update global stats & pbar
        global_pbar.update(1)
        global_pbar.set_postfix_str(f"✅ {global_stats['ok']} | ⚠️ {global_stats['partial']} | ❌ {global_stats['fail']}")
        
    # Ringkasan Akhir Hari (Pake emot di tiap kategori biar makin jelas)
    global_pbar.write(f"\n📊 Summary {tanggal_str}: ✅ \033[1;32m{ok_count} OK\033[0m | ⚠️  \033[1;33m{na_count} Partial\033[0m | ❌ \033[1;31m{fail_count} Fail\033[0m")
    global_pbar.write(f"{dim}──────────────────────────────────────────────────{reset}\n")


# ========================================================
#  MODE 2: IMAGE ONLY (insert gambar saja, tanpa OCR)
# ========================================================

def baca_mapping_img(filepath):
    """Baca file mapping Image-only (format SID -> range)."""
    mapping = {}
    with open(filepath, 'r', encoding='utf-8') as f:
        lines = [line.strip() for line in f if line.strip()]
    i = 0
    while i < len(lines):
        line = lines[i]
        if line.startswith('SID : '):
            id_raw = line.replace('SID : ', '').strip()
            id_clean = re.sub(r'^\(\d+\)\s*', '', id_raw)
            i += 1
            if i < len(lines) and lines[i].startswith('->'):
                range_str = lines[i][2:].strip()
                start, end = range_str.split('-')
                start_col = column_index_from_string(re.match(r'[A-Z]+', start).group())
                start_row = int(re.search(r'\d+', start).group())
                end_col = column_index_from_string(re.match(r'[A-Z]+', end).group())
                end_row = int(re.search(r'\d+', end).group())
                mapping[id_clean] = ((start_row, start_col), (end_row, end_col))
                i += 1
            else:
                i += 1
        else:
            i += 1
    return mapping


def proses_tanggal_img(wb, tanggal_str, items, mapping, tanggal_idx, total_tanggal):
    """Proses satu folder tanggal untuk mode Image-only (insert gambar saja)."""
    hari = int(tanggal_str[6:8])
    sheet_name = f"{hari:02d}"
    if sheet_name not in wb.sheetnames:
        # Fallback: coba tanpa leading zero
        sheet_name = str(hari)
        if sheet_name not in wb.sheetnames:
            console.print(f"[red]❌[/red] Sheet {hari:02d} tidak ditemukan")
            return
    sheet = wb[sheet_name]

    pbar = tqdm(items, desc=f"Tanggal {tanggal_str} ({tanggal_idx}/{total_tanggal})", 
                leave=True, ncols=100, bar_format='{l_bar}{bar}| {n_fmt}/{total_fmt}',
                file=sys.stdout)

    for nomor, tipe, id_val in pbar:
        id_clean = re.sub(r'^\(\d+\)\s*', '', id_val)
        if id_clean not in mapping:
            pbar.write(f"  ⚠️  Peringatan: ID '{id_clean}' tidak ditemukan di mapping")
            continue
        (start_row, start_col), (end_row, end_col) = mapping[id_clean]

        path_gambar = cari_path_gambar(FOLDER_DATA, tanggal_str, tipe, id_val)
        if not os.path.exists(path_gambar):
            pbar.write(f"  ❌ Gambar tidak ditemukan: {path_gambar}")
            continue

        pbar.write(f"  ✅ Menambahkan gambar untuk {tipe} {id_val}")
        tambah_gambar_di_area(sheet, path_gambar, start_row, start_col, end_row, end_col)

    pbar.close()


# ========================================================
#  MAIN
# ========================================================

def main():
    console.print("\n" + "=" * TERMINAL_WIDTH, style="bold cyan")
    console.print("  AUTOMATED MRTG TO EXCEL REPORT", style="bold cyan")
    console.print("=" * TERMINAL_WIDTH + "\n", style="bold cyan")
    console.print("  Pilih mode:")
    console.print("  [1] OCR Mode   : Ekstrak data + insert gambar ke Excel")
    console.print("  [2] Image Only : Insert gambar saja ke Excel (tanpa OCR)")
    console.print("=" * TERMINAL_WIDTH)

    while True:
        pilihan = input("\n  >> Masukkan pilihan (1/2): ").strip()
        if pilihan in ('1', '2'):
            break
        console.print("[red]Input tidak valid. Masukkan 1 atau 2.[/red]")

    if pilihan == '1':
        # === MODE OCR ===
        console.print("\n[green]✅[/green] Mode: [bold]OCR[/bold] (Ekstrak data + insert gambar)\n", style="bold green")
        template_file = OCR_TEMPLATE
        output_file   = OCR_OUTPUT
        mapping_file  = OCR_MAPPING
        daftar_file   = OCR_DAFTAR

        if not os.path.exists(mapping_file):
            console.print(f"[red]❌ File mapping '{mapping_file}' tidak ditemukan![/red]")
            return
        mapping = baca_mapping_ocr(mapping_file)
        console.print(f"[cyan]📋[/cyan] Mapping berisi [bold]{len(mapping)}[/bold] entri.")

        items = baca_daftar(daftar_file)
        console.print(f"[cyan]📝[/cyan] Daftar berisi [bold]{len(items)}[/bold] item.")

        tanggal_list = get_tanggal_list(FOLDER_DATA)
        if not tanggal_list:
            console.print(f"[red]❌ Folder '{FOLDER_DATA}' tidak ditemukan atau kosong![/red]")
            return
        console.print(f"[cyan]📅[/cyan] Ditemukan [bold]{len(tanggal_list)}[/bold] folder tanggal.")

        if not os.path.exists(template_file):
            console.print(f"[red]❌ Template '{template_file}' tidak ditemukan![/red]")
            return
        wb = load_workbook(template_file)
        console.print("[cyan]✅[/cyan] Template berhasil dimuat.\n")

        # LOG SESSION START
        logger.info("")
        logger.info("="*80)
        logger.info(f"🆕 SESSION START: MODE OCR")
        logger.info(f"📂 Template: {os.path.basename(template_file)}")
        logger.info("="*80)

        # Global Stats
        global_stats = {
            'ok': 0, 'partial': 0, 'fail': 0, 
            'total_items': len(tanggal_list) * len(items)
        }
        review_list = []

        # 1. PAKSA INISIALISASI OCR DI SINI
        console.print("[cyan]⚙️  Initializing OCR Engine...[/cyan]")
        import contextlib
        with open(os.devnull, 'w') as fnull:
            with contextlib.redirect_stdout(fnull), contextlib.redirect_stderr(fnull):
                _get_ocr_engine()
        console.print("[green]✅ Engine Ready![/green]\n")

        # 2. Progress bar STICKY di paling bawah
        global_pbar = tqdm(total=global_stats['total_items'], desc="Progres", ncols=PBAR_WIDTH_OCR, 
                           bar_format='{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}, {rate_fmt}] {postfix}',
                           file=sys.stdout)

        try:
            for tgl_idx, tgl in enumerate(tanggal_list, 1):
                proses_tanggal_ocr(wb, tgl, items, mapping, global_stats, review_list, tgl_idx, len(tanggal_list), global_pbar)
            global_pbar.close()

            # LOG SESSION END
            logger.info("="*80)
            logger.info(f"🏁 SESSION END: SUCCESS")
            logger.info(f"📊 Summary: OK={global_stats['ok']}, Partial={global_stats['partial']}, Fail={global_stats['fail']}")
            logger.info("="*80)
            logger.info("")
        except Exception as e:
            console.print(f"\n[bold red]❌ Terjadi kesalahan fatal saat memproses:[/bold red] {e}")
            logger.error(f"Fatal Loop Error: {e}\n{traceback.format_exc()}")
            logger.info("╔" + "═"*78 + "╗")
            logger.info(f"║ 🏁 SESSION END: ERROR ({str(e)[:56]:<57} ║")
            logger.info("╚" + "═"*78 + "╝\n")
            
        wb.save(output_file)
        
        # --- FINAL SUMMARY ---
        console.print("\n" + "=" * TERMINAL_WIDTH, style="bold cyan")
        console.print("  FINAL REPORT SUMMARY", style="bold cyan")
        console.print("=" * TERMINAL_WIDTH + "\n", style="bold cyan")
        
        total = global_stats['total_items']
        if total > 0:
            ok_p = (global_stats['ok'] / total * 100)
            partial_p = (global_stats['partial'] / total * 100)
            fail_p = (global_stats['fail'] / total * 100)
            
            console.print(f"  [green]✅[/green] BERHASIL (100%)    : [bold green]{global_stats['ok']}[/bold green] items ({ok_p:.1f}%)")
            console.print(f"  [yellow]⚠️ [/yellow] PARTIAL (N/A)     : [bold yellow]{global_stats['partial']}[/bold yellow] items ({partial_p:.1f}%)")
            console.print(f"  [red]❌[/red] GAGAL (No Data)    : [bold red]{global_stats['fail']}[/bold red] items ({fail_p:.1f}%)")
            console.print(f"  [cyan]─[/cyan] " + "─"*56)
            console.print(f"  [cyan]📊[/cyan] TOTAL ITEM PROSES : [bold cyan]{total}[/bold cyan] items")
        else:
            console.print("  Tidak ada item yang diproses.", style="yellow")
        console.print("="*60)

        if review_list:
            console.print("\n[yellow]🔍[/yellow] [bold]LIST ITEM YANG PERLU DICEK[/bold] (GROUPED BY SID):\n", style="bold yellow")
            
            # Grouping by SID
            grouped_issues = {}
            for item in review_list:
                sid = item['sid']
                if sid not in grouped_issues:
                    grouped_issues[sid] = {
                        'dates': [],
                        'sheets': [],
                        'total_na': 0,
                        'partial_count': 0,
                        'fail_count': 0
                    }
                
                grouped_issues[sid]['dates'].append(item['date'])
                grouped_issues[sid]['sheets'].append(item['sheet'])
                grouped_issues[sid]['total_na'] += item['na']
                if item['status'] == 'Partial':
                    grouped_issues[sid]['partial_count'] += 1
                else:
                    grouped_issues[sid]['fail_count'] += 1

            for sid, data in grouped_issues.items():
                dates_str = ", ".join(sorted(list(set(data['dates']))))
                sheets_str = ", ".join(sorted(list(set(data['sheets'])), key=lambda x: int(x) if x.isdigit() else 0))
                
                status_parts = []
                if data['partial_count'] > 0:
                    status_parts.append(f"[yellow]{data['partial_count']} partial[/yellow]")
                if data['fail_count'] > 0:
                    status_parts.append(f"[red]{data['fail_count']} fail[/red]")
                status_str = " & ".join(status_parts)
                
                console.print(f"[cyan]SID[/cyan]     : [bold]{sid}[/bold]")
                console.print(f"[cyan]Tanggal[/cyan] : {dates_str}")
                console.print(f"[cyan]Sheet[/cyan]   : {sheets_str}")
                console.print(f"[cyan]Status[/cyan]  : {status_str}, [bold]{data['total_na']}[/bold] total nilai N/A")
                console.print("[cyan]" + "─"*56 + "[/cyan]")
            
            console.print(f"\n[yellow]ℹ️ [/yellow] Total [bold yellow]{len(grouped_issues)}[/bold yellow] SID unik butuh review.", style="bold yellow")
            console.print("[cyan]💡[/cyan] Tips: Cek '[bold]ocr_report.log[/bold]' untuk melihat detail teks yang terdeteksi.", style="cyan")
        else:
            console.print("\n[green]✨[/green] [bold green]SEMPURNA![/bold green] Semua data terisi 100%. Tidak ada item untuk direview.", style="bold green")

        console.print(f"\n[cyan]📁[/cyan] File output: [bold blue]{os.path.abspath(output_file)}[/bold blue]")
        console.print("="*60)

    else:
        # === MODE IMAGE ONLY ===
        console.print("\n[green]✅[/green] Mode: [bold]Image Only[/bold] (Insert gambar saja)\n", style="bold green")
        template_file = IMG_TEMPLATE
        output_file   = IMG_OUTPUT
        mapping_file  = IMG_MAPPING
        daftar_file   = IMG_DAFTAR

        if not os.path.exists(mapping_file):
            console.print(f"[red]❌ File mapping '{mapping_file}' tidak ditemukan![/red]")
            return
        mapping = baca_mapping_img(mapping_file)
        console.print(f"[cyan]📋[/cyan] Mapping berisi [bold]{len(mapping)}[/bold] entri.")

        items = baca_daftar(daftar_file)
        console.print(f"[cyan]📝[/cyan] Daftar berisi [bold]{len(items)}[/bold] item.")

        tanggal_list = get_tanggal_list(FOLDER_DATA)
        if not tanggal_list:
            console.print(f"[red]❌ Folder '{FOLDER_DATA}' tidak ditemukan atau kosong![/red]")
            return
        console.print(f"[cyan]📅[/cyan] Ditemukan [bold]{len(tanggal_list)}[/bold] folder tanggal.")

        if not os.path.exists(template_file):
            console.print(f"[red]❌ Template '{template_file}' tidak ditemukan![/red]")
            return
        wb = load_workbook(template_file)
        console.print("[cyan]✅[/cyan] Template berhasil dimuat.\n")

        # LOG SESSION START (IMG ONLY)
        logger.info("")
        logger.info("="*80)
        logger.info(f"🆕 SESSION START: MODE IMAGE ONLY")
        logger.info(f"📂 Template: {os.path.basename(template_file)}")
        logger.info("="*80)

        try:
            for tgl_idx, tgl in enumerate(tqdm(tanggal_list, desc="Total Progres", ncols=PBAR_WIDTH_IMG, file=sys.stdout), 1):
                proses_tanggal_img(wb, tgl, items, mapping, tgl_idx, len(tanggal_list))
            
            # LOG SESSION END (IMG ONLY)
            logger.info("="*80)
            logger.info(f"🏁 SESSION END: SUCCESS (IMAGE ONLY)")
            logger.info("="*80)
            logger.info("")
        except Exception as e:
            console.print(f"\n[bold red]❌ Terjadi kesalahan fatal:[/bold red] {e}")
            logger.error(f"Fatal Loop Error: {e}")
            logger.info("╔" + "═"*78 + "╗")
            logger.info(f"║ 🏁 SESSION END: ERROR ({str(e)[:56]:<57} ║")
            logger.info("╚" + "═"*78 + "╝\n")
            
        wb.save(output_file)
        console.print("\n" + "="*60, style="bold cyan")
        console.print(f"  [green]🎉[/green] [bold]SELESAI![/bold] File Excel disimpan sebagai:", style="bold green")
        console.print(f"  [bold blue]{os.path.abspath(output_file)}[/bold blue]", style="bold")
        console.print("="*60)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        console.print("\n\n[bold yellow]⚠️  Proses dihentikan paksa oleh pengguna (Ctrl+C).[/bold yellow]")
        sys.exit(0)
    except Exception as e:
        console.print(f"\n\n[bold red]❌ Terjadi error tak terduga:[/bold red] {e}")
        sys.exit(1)